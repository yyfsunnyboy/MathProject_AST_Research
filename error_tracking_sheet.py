import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_error_tracking_sheet(workbook):
    """Create a new worksheet for tracking errors with AI explanations and guidance"""
    
    # Create new worksheet
    ws = workbook.create_sheet("Error Tracking")
    
    # Define styles
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define columns
    columns = [
        "序號",
        "日期時間",
        "題目",
        "用戶答案",
        "正確答案",
        "錯誤種類",
        "AI簡短說明",
        "AI提供指導",
        "難度等級"
    ]
    
    # Add headers
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    # Set column widths
    column_widths = [8, 18, 30, 15, 15, 15, 25, 30, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    return ws

# Usage example
def add_error_record(ws, error_data):
    """Add a new error record to the Error Tracking worksheet"""
    
    row = ws.max_row + 1
    
    record = [
        error_data.get("序號"),
        error_data.get("日期時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        error_data.get("題目"),
        error_data.get("用戶答案"),
        error_data.get("正確答案"),
        error_data.get("錯誤種類"),
        error_data.get("AI簡短說明"),
        error_data.get("AI提供指導"),
        error_data.get("難度等級")
    ]
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, value in enumerate(record, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    ws.row_dimensions[row].height = 25